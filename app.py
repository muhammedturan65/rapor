import io
import re
import os
import sys
from datetime import datetime
from flask import Flask, request, send_file, render_template_string
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

app = Flask(__name__)

# --- GÖRSEL ARAYÜZ ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="tr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Happy Center Rapor Sistemi</title>
    <style>
        body { font-family: 'Calibri', sans-serif; background-color: #f4f4f9; display: flex; justify-content: center; align-items: center; height: 100vh; margin: 0; }
        .container { background: white; padding: 40px; border-radius: 10px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); text-align: center; width: 450px; }
        h1 { color: #e67e22; margin: 0; font-size: 32px; }
        h2 { color: #555; margin-top: 5px; font-size: 18px; font-weight: normal;}
        .info { background-color: #e8f8f5; color: #0e6251; padding: 10px; border-radius: 5px; font-size: 13px; margin-bottom: 20px; border: 1px solid #a2d9ce;}
        .upload-btn-wrapper { position: relative; overflow: hidden; display: inline-block; margin-bottom: 20px; width: 100%; }
        .btn { border: 2px dashed #e67e22; color: #e67e22; background-color: white; padding: 20px; border-radius: 8px; font-size: 16px; font-weight: bold; cursor: pointer; transition: 0.3s; width: 100%; box-sizing: border-box;}
        .btn:hover { background-color: #fdf2e9; }
        input[type=file] { font-size: 100px; position: absolute; left: 0; top: 0; opacity: 0; cursor: pointer; height: 100%; width: 100%; }
        .file-name { margin-top: -10px; margin-bottom: 15px; color: #333; font-weight: bold; font-size: 14px; display: block; min-height: 20px;}
        .convert-btn { background-color: #27ae60; color: white; border: none; padding: 15px; width: 100%; border-radius: 8px; font-size: 18px; font-weight: bold; cursor: pointer; box-shadow: 0 4px #1e8449; transition: 0.1s; }
        .convert-btn:active { box-shadow: 0 2px #1e8449; transform: translateY(2px); }
        .convert-btn:hover { background-color: #2ecc71; }
    </style>
</head>
<body>
    <div class="container">
        <h1>HAPPY CENTER</h1>
        <h2>Şube Açılış/Kapanış Raporu</h2>
        
        <div class="info">
            Özellikler: Akıllı Eşleştirme + Depo Renklendirme<br>
            (Merkez Depolar mavi dolgulu, Sıra No kırmızı)
        </div>

        <form action="/" method="post" enctype="multipart/form-data">
            <div class="upload-btn-wrapper">
                <div class="btn">📂 Dosya Seçmek İçin Tıklayın</div>
                <input type="file" name="file" required onchange="document.getElementById('fname').innerText = this.files[0].name; document.querySelector('.btn').style.borderColor='#27ae60'; document.querySelector('.btn').style.color='#27ae60';">
            </div>
            <span id="fname" class="file-name"></span>
            
            <button type="submit" class="convert-btn">Excel'e Dönüştür ve İndir</button>
        </form>
    </div>
</body>
</html>
"""

# --- YARDIMCI FONKSİYONLAR ---
def tarihi_formatla(tarih_metni):
    try:
        aylar = {
            "Ocak": "01", "Şubat": "02", "Mart": "03", "Nisan": "04", "Mayıs": "05", "Haziran": "06",
            "Temmuz": "07", "Ağustos": "08", "Eylül": "09", "Ekim": "10", "Kasım": "11", "Aralık": "12"
        }
        temiz_metin = tarih_metni.replace(":", "").strip()
        parcalar = temiz_metin.split()
        if len(parcalar) >= 3:
            gun = parcalar[0].zfill(2)
            ay_isim = parcalar[1]
            yil = parcalar[2]
            ay_no = aylar.get(ay_isim, "00")
            if ay_no != "00":
                return f"{gun}.{ay_no}.{yil}"
        return temiz_metin
    except:
        return tarih_metni

def normalize_tr(text):
    if not text: return ""
    tr_map = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")
    text = text.translate(tr_map).upper()
    return re.sub(r'[^A-Z0-9]', '', text)

def find_file(filename):
    possible_paths = [
        filename,
        os.path.join(os.getcwd(), filename),
        os.path.join(os.path.dirname(__file__), filename),
        f'/var/task/{filename}'
    ]
    for path in possible_paths:
        if os.path.exists(path):
            return path
    return None

def get_sorting_data():
    path = find_file('siralama.txt')
    ordered_list = []
    norm_map = {} 
    
    if path:
        try:
            with open(path, 'r', encoding='utf-8') as f:
                for line in f:
                    orijinal_isim = line.strip()
                    if orijinal_isim:
                        ordered_list.append(orijinal_isim)
                        norm_key = normalize_tr(orijinal_isim)
                        norm_map[norm_key] = orijinal_isim
        except Exception as e:
            print(f"Sıralama dosyası okunamadı: {e}")
    return ordered_list, norm_map

def akilli_eslestir(html_name, norm_map):
    h_norm = normalize_tr(html_name)
    
    # Senin belirlediğin özel düzeltmeler
    manual_fixes = {
        "BESYUZEVLER1": "BESYUZEVLER",
        "KUCUKCEKMECE": "CEKMECE",
        "HALKALI1": "HALKALI",
        "CUMHURIYET": "CUMHURRIYET",
        "BEYKENT1": "BEYKENT",
        "ZUMRUTEVLER1": "ZUMRUTEVELER",
        "ORNEKMAH1": "ORNEKMAHALLESI",
        "EYUPISLAMBEY": "EYUP2",
        "ORNEKMAH2": "ORNEKMAHALLESI2",
        "YENIETICARET": "YENIETICARETDEPO",
        "IHLASMARMARA1": "ILHASMARMARA1",
        "PARKMAVERA": "PARKMEVRA",
        "UGURMUMCUADNANKHVCI": "UGURMUMCUADNANK",
        "KURTKOYMILLETCAD": "KURTKOYMILLETCADDESI",
        "BAHCELIEVLERYAYLA": "BEHCELIEVLERYAYLA",
        "PASABAYIR": "PASABAYIR1",
        "BIGA1": "BIGA",
        "KARACABEY": "KARACABEY1",
        "BALIKESIR": "BALIKESIR1",
        "BANDIRMAMERKEZ": "BADIRMAMERKEZ",
        "BALIKESIRRAMADA": "BALIKKESIRRAMADA",
        "GONENKURTULUS": "GONENKUTULUS",
        "IKBALALTUN": "IKBALALTUN"
    }
    
    if h_norm in manual_fixes:
        target_norm = manual_fixes[h_norm]
        if target_norm in norm_map:
            return norm_map[target_norm]

    if h_norm in norm_map:
        return norm_map[h_norm]
    
    for t_norm, t_original in norm_map.items():
        if t_norm in h_norm and len(t_norm) > 3: 
            return t_original
        if h_norm in t_norm and len(h_norm) > 3:
            return t_original

    return html_name

def process_html_to_excel(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    satirlar = soup.find_all('tr')
    
    custom_order_list, norm_map = get_sorting_data()
    
    sube_verileri = {}
    aktif_sube = None
    rapor_tarihi_html = None

    # HTML Veri Çekme
    for satir in satirlar:
        satir_metni = satir.get_text(" ", strip=True)
        hucreler = satir.find_all('td')

        if "Rapor Tarihi" in satir_metni:
            for td in hucreler:
                txt = td.get_text(strip=True)
                if txt and "Rapor Tarihi" not in txt:
                    rapor_tarihi_html = tarihi_formatla(txt)
                    break
            continue

        if "Firma Ünvanı" in satir_metni:
            for td in hucreler:
                txt = td.get_text(strip=True)
                if txt and "Firma Ünvanı" not in txt and len(txt) > 2:
                    aktif_sube = txt
                    break
            continue
        
        if not "Firma Ünvanı" in satir_metni:
            genis_hucre = satir.find('td', attrs={'colspan': '28'})
            if genis_hucre:
                txt = genis_hucre.get_text(strip=True)
                if txt: aktif_sube = txt

        if aktif_sube and ("SİSTEM KAPATILDI" in satir_metni or "SİSTEM KURULDU" in satir_metni or "MAĞAZA AÇILDI" in satir_metni or "MAĞAZA KAPANDI" in satir_metni):
            duzeltilmis_isim = akilli_eslestir(aktif_sube, norm_map)

            if duzeltilmis_isim not in sube_verileri:
                sube_verileri[duzeltilmis_isim] = {"acilis_saat": "", "acilis_kisi": "", "kapanis_saat": "", "kapanis_kisi": ""}

            saat_match = re.search(r'\b([0-9]{1,2}):([0-9]{2})\b', satir_metni)
            saat = saat_match.group(0) if saat_match else ""

            ham_veri = ""
            for td in hucreler:
                txt_td = td.get_text(strip=True)
                if "SİSTEM" in txt_td or "MAĞAZA" in txt_td:
                    ham_veri = txt_td
                    break
            if not ham_veri: ham_veri = satir_metni
            
            personel = ""
            separator = ""
            if "SİSTEM" in ham_veri: separator = "SİSTEM"
            elif "MAĞAZA" in ham_veri: separator = "MAĞAZA"

            if separator:
                personel = ham_veri.split(separator)[0].strip(" .")
                personel = re.sub(r'^\d+\.\s*', '', personel)

            if "SİSTEM KAPATILDI" in satir_metni or "MAĞAZA AÇILDI" in satir_metni:
                # İlk açılışı al: Eğer saat henüz kaydedilmemişse kaydet
                if not sube_verileri[duzeltilmis_isim]["acilis_saat"]:
                    sube_verileri[duzeltilmis_isim]["acilis_saat"] = saat
                    sube_verileri[duzeltilmis_isim]["acilis_kisi"] = personel
            elif "SİSTEM KURULDU" in satir_metni or "MAĞAZA KAPANDI" in satir_metni:
                # Son kapanışı al: Her seferinde üzerine yaz, böylece en sonuncusu kalır
                sube_verileri[duzeltilmis_isim]["kapanis_saat"] = saat
                sube_verileri[duzeltilmis_isim]["kapanis_kisi"] = personel

    # --- LİSTE OLUŞTURMA ---
    final_sorted_keys = []
    if custom_order_list:
        final_sorted_keys = list(custom_order_list)
    
    existing_norms = {normalize_tr(k) for k in final_sorted_keys}
    for sube_adi in sube_verileri:
        if normalize_tr(sube_adi) not in existing_norms:
            final_sorted_keys.append(sube_adi)
            
    if not final_sorted_keys:
        final_sorted_keys = sorted(sube_verileri.keys())

    # --- EXCEL TASARIMI ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Happy Center Rapor"

    # FONTLAR VE STİLLER
    font_main_title = Font(name='Calibri', size=14, bold=True)
    font_header = Font(name='Calibri', size=12, bold=True)
    font_branch = Font(name='Calibri', size=16, bold=True)
    font_normal = Font(name='Calibri', size=12)
    font_acilis = Font(name='Calibri', size=12, color="006100")
    font_kapanis = Font(name='Calibri', size=12, color="000080")
    
    # YENİ EKLENEN STİLLER
    font_sira_no = Font(name='Calibri', size=12, color="FF0000") # Kırmızı Sıra No
    blue_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid") # Merkez Depo Mavisi

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    # Logo
    ws.row_dimensions[1].height = 48.75
    logo_path = find_file('logo.png')
    if logo_path:
        try:
            img1 = XLImage(logo_path); img1.width = 264; img1.height = 65
            ws.add_image(img1, 'A1')
            img2 = XLImage(logo_path); img2.width = 264; img2.height = 65
            ws.add_image(img2, 'G1')
        except: pass

    # Başlık
    ws.merge_cells('B1:F1')
    ws['B1'] = "HAPPY CENTER ŞUBELERİN AÇILIŞ KAPANIŞ SAATLERİ"
    ws['B1'].font = font_main_title
    ws['B1'].alignment = center_align

    # Tarih
    tarih_str = rapor_tarihi_html if rapor_tarihi_html else datetime.now().strftime("%d.%m.%Y")
    ws.merge_cells('B2:F2')
    ws['B2'] = f"{tarih_str} HAPPY CENTER MAĞAZA AÇILIŞ VE KAPANIŞLARI"
    ws['B2'].font = font_header
    ws['B2'].alignment = center_align
    ws.row_dimensions[2].height = 26.25 

    headers_config = [
        ('A4:A5', 'SIRA NO'), ('B4:B5', 'ŞUBE ADI'),
        ('C4:D4', 'ŞUBEYİ AÇAN'), ('C5', 'SAAT'), ('D5', 'PERSONEL'),
        ('E4:F4', 'ŞUBEYİ KAPATAN'), ('E5', 'SAAT'), ('F5', 'PERSONEL'),
        ('G4:G5', 'AÇIKLAMA')
    ]
    for rng, val in headers_config:
        if ':' in rng: ws.merge_cells(rng)
        cell = ws[rng.split(':')[0]]
        cell.value = val; cell.fill = header_fill; cell.font = font_header
        cell.alignment = center_align; cell.border = thin_border
    
    ws.row_dimensions[4].height = 15.75
    ws.row_dimensions[5].height = 15.75
    for r in ws['A4:G5']:
        for c in r: c.border = thin_border

    # Verileri Yaz
    start_row = 6
    sira_no = 1
    
    # Özel renklendirilecek şubeler listesi
    ozel_subeler = ["MERKEZ DEPO", "MERKEZ DEPO MUTFAK", "MERKEZ DEPO_2"]
    
    for sube in final_sorted_keys:
        data = sube_verileri.get(sube, {
            "acilis_saat": "", "acilis_kisi": "",
            "kapanis_saat": "", "kapanis_kisi": ""
        })
        
        ws.row_dimensions[start_row].height = 15.75
        
        # Sütun 1: Sıra No (KIRMIZI FONT)
        c = ws.cell(row=start_row, column=1, value=sira_no)
        c.font = font_sira_no 
        c.border = thin_border
        c.alignment = center_align
        
        # Sütun 2: Şube Adı
        c = ws.cell(row=start_row, column=2, value=sube)
        c.font = font_branch
        c.border = thin_border
        c.alignment = Alignment(vertical='center', indent=1)

        # Bu satır özel şube mi?
        is_merkez = sube in ozel_subeler

        # Sütun 3: Açılış Saat (Özel ise Mavi Dolgu)
        c = ws.cell(row=start_row, column=3, value=data['acilis_saat'])
        c.font = font_acilis
        c.alignment = center_align
        c.border = thin_border
        if is_merkez: c.fill = blue_fill # Mavi Dolgu

        # Sütun 4: Açılış Personel
        c = ws.cell(row=start_row, column=4, value=data['acilis_kisi'])
        c.font = font_acilis
        c.border = thin_border
        c.alignment = Alignment(vertical='center', indent=1)

        # Sütun 5: Kapanış Saat (Özel ise Mavi Dolgu)
        c = ws.cell(row=start_row, column=5, value=data['kapanis_saat'])
        c.font = font_kapanis
        c.alignment = center_align
        c.border = thin_border
        if is_merkez: c.fill = blue_fill # Mavi Dolgu
        
        # Sütun 6: Kapanış Personel
        c = ws.cell(row=start_row, column=6, value=data['kapanis_kisi'])
        c.font = font_kapanis
        c.border = thin_border
        c.alignment = Alignment(vertical='center', indent=1)
        
        # Sütun 7: Açıklama
        c = ws.cell(row=start_row, column=7, value="")
        c.border = thin_border

        start_row += 1
        sira_no += 1

    ws.column_dimensions['A'].width = 8.43
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 9.14
    ws.column_dimensions['D'].width = 29
    ws.column_dimensions['E'].width = 9.14
    ws.column_dimensions['F'].width = 32
    ws.column_dimensions['G'].width = 68.86

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            return "Lütfen bir dosya seçin!", 400
        try:
            html_content = file.read().decode('utf-8')
            excel_file = process_html_to_excel(html_content)
            dosya_adi = f"Happy_Center_Rapor_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
            return send_file(excel_file, as_attachment=True, download_name=dosya_adi, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            return f"Bir hata oluştu: {str(e)}", 500
    return render_template_string(HTML_TEMPLATE)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
